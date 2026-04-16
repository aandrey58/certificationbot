from PyQt5 import QtCore, QtGui, QtWidgets

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import icon_rc  # ресурс с иконкой (:/icon.ico)


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.setFixedSize(820, 260)
        MainWindow.setWindowFlag(QtCore.Qt.WindowMaximizeButtonHint, False)

        # Иконка окна из ресурсного файла (resources.qrc / icon_rc.py)
        MainWindow.setWindowIcon(QtGui.QIcon(":/icon.ico"))

        # Общий стиль окна
        MainWindow.setStyleSheet(
            """
            QWidget {
                font-family: "Segoe UI";
                font-size: 10pt;
            }
            QLineEdit {
                padding: 4px 6px;
                border: 1px solid #d0d0d0;
                border-radius: 4px;
                background: #ffffff;
            }
            QLineEdit:focus {
                border: 1px solid #0078d7;
            }
            QLabel#titleLabel {
                font-size: 14pt;
                font-weight: 600;
            }
            QLabel#hintLabel {
                color: #666666;
            }
            QPushButton {
                min-height: 28px;
                padding: 4px 18px;
                border-radius: 4px;
                border: 1px solid #0078d7;
                background-color: #0078d7;
                color: #ffffff;
            }
            QPushButton:disabled {
                background-color: #a0c6ea;
                border-color: #a0c6ea;
                color: #f5f5f5;
            }
            QPushButton#secondaryButton {
                background-color: #f3f3f3;
                color: #333333;
                border: 1px solid #c4c4c4;
            }
            QPushButton#secondaryButton:hover {
                background-color: #e5e5e5;
            }
            QPushButton:hover {
                background-color: #1a82e2;
            }
            QToolButton {
                border: 1px solid #d0d0d0;
                border-radius: 4px;
                padding: 0 6px;
                background-color: #f5f5f5;
            }
            QToolButton:hover {
                background-color: #e5e5e5;
            }
            """
        )

        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        mainLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        mainLayout.setContentsMargins(16, 12, 16, 12)
        mainLayout.setSpacing(10)

        # Заголовок
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setObjectName("label_4")
        self.label_4.setAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
        self.label_4.setProperty("class", "title")
        self.label_4.setObjectName("titleLabel")
        mainLayout.addWidget(self.label_4)

        formLayout = QtWidgets.QGridLayout()
        formLayout.setColumnStretch(1, 1)
        formLayout.setHorizontalSpacing(8)
        formLayout.setVerticalSpacing(6)

        # Строка выбора Excel
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setObjectName("label")
        formLayout.addWidget(self.label, 0, 0, 1, 1)

        excelLayout = QtWidgets.QHBoxLayout()
        excelLayout.setContentsMargins(0, 0, 0, 0)
        excelLayout.setSpacing(4)
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setObjectName("lineEdit")
        excelLayout.addWidget(self.lineEdit)
        self.toolButton = QtWidgets.QToolButton(self.centralwidget)
        self.toolButton.setObjectName("toolButton")
        excelLayout.addWidget(self.toolButton)
        formLayout.addLayout(excelLayout, 0, 1, 1, 1)

        # Строка выбора JSON
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setObjectName("label_2")
        formLayout.addWidget(self.label_2, 1, 0, 1, 1)

        jsonLayout = QtWidgets.QHBoxLayout()
        jsonLayout.setContentsMargins(0, 0, 0, 0)
        jsonLayout.setSpacing(4)
        self.lineEdit_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_2.setObjectName("lineEdit_2")
        jsonLayout.addWidget(self.lineEdit_2)
        self.toolButton_2 = QtWidgets.QToolButton(self.centralwidget)
        self.toolButton_2.setObjectName("toolButton_2")
        jsonLayout.addWidget(self.toolButton_2)
        formLayout.addLayout(jsonLayout, 1, 1, 1, 1)

        # Строка имени листа
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setObjectName("label_3")
        formLayout.addWidget(self.label_3, 2, 0, 1, 1)

        sheetLayout = QtWidgets.QHBoxLayout()
        sheetLayout.setContentsMargins(0, 0, 0, 0)
        sheetLayout.setSpacing(4)
        self.lineEdit_4 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_4.setObjectName("lineEdit_4")
        sheetLayout.addWidget(self.lineEdit_4)
        formLayout.addLayout(sheetLayout, 2, 1, 1, 1)

        mainLayout.addLayout(formLayout)

        # Подсказка
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setObjectName("hintLabel")
        self.label_5.setWordWrap(True)
        mainLayout.addWidget(self.label_5)

        # Кнопки внизу
        buttonsLayout = QtWidgets.QHBoxLayout()
        buttonsLayout.addStretch(1)

        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setObjectName("pushButton")
        buttonsLayout.addWidget(self.pushButton)

        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setObjectName("secondaryButton")
        buttonsLayout.addWidget(self.pushButton_2)

        mainLayout.addLayout(buttonsLayout)

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 785, 26))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        # --- connections ---
        self.toolButton.clicked.connect(self.choose_excel_file)
        self.toolButton_2.clicked.connect(self.choose_json_file)
        self.pushButton.clicked.connect(self.create_sheet_from_json)
        self.pushButton_2.clicked.connect(MainWindow.close)

        self._main_window = MainWindow

        # Изначально кнопка "Создать" неактивна, пока поля не заполнены
        self.pushButton.setEnabled(False)
        self.lineEdit.textChanged.connect(self.update_create_button_state)
        self.lineEdit_2.textChanged.connect(self.update_create_button_state)
        self.lineEdit_4.textChanged.connect(self.update_create_button_state)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Заполнение Excel-файла"))
        self.pushButton.setText(_translate("MainWindow", "Создать"))
        self.pushButton_2.setText(_translate("MainWindow", "Закрыть"))
        self.label.setText(_translate("MainWindow", "Выберите целевой Excel-файл:"))
        self.toolButton.setText(_translate("MainWindow", "..."))
        self.toolButton_2.setText(_translate("MainWindow", "..."))
        self.label_2.setText(_translate("MainWindow", "Выберите json-файл:"))
        self.label_3.setText(_translate("MainWindow", "Введите название листа*:"))
        self.label_4.setText(_translate("MainWindow", "Заполнение Excel-файла"))
        self.label_5.setText(_translate("MainWindow", "*название листа отображается в боте как название области, например, Б 2.2"))

    # ---------- UI helpers ----------

    def choose_excel_file(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self._main_window,
            "Выберите Excel-файл",
            "",
            "Excel (*.xlsx)",
        )
        if path:
            self.lineEdit.setText(path)

    def choose_json_file(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self._main_window,
            "Выберите JSON-файл",
            "",
            "JSON (*.json)",
        )
        if path:
            self.lineEdit_2.setText(path)

    def _show_error(self, message: str):
        QtWidgets.QMessageBox.critical(self._main_window, "Ошибка", message)

    def _show_info(self, message: str):
        QtWidgets.QMessageBox.information(self._main_window, "Готово", message)

    def update_create_button_state(self):
        has_excel = bool(self.lineEdit.text().strip())
        has_json = bool(self.lineEdit_2.text().strip())
        has_sheet = bool(self.lineEdit_4.text().strip())
        self.pushButton.setEnabled(has_excel and has_json and has_sheet)

    # ---------- Core logic ----------

    def create_sheet_from_json(self):
        excel_path = self.lineEdit.text().strip()
        json_path = self.lineEdit_2.text().strip()
        sheet_name = self.lineEdit_4.text().strip()

        if not excel_path:
            self._show_error("Не указан Excel-файл.")
            return
        if not json_path:
            self._show_error("Не указан JSON-файл.")
            return
        if not sheet_name:
            self._show_error("Не указано название листа.")
            return

        excel_file = Path(excel_path)
        json_file = Path(json_path)

        if not excel_file.is_file():
            self._show_error("Указанный Excel-файл не существует.")
            return
        if not json_file.is_file():
            self._show_error("Указанный JSON-файл не существует.")
            return

        try:
            with json_file.open("r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            self._show_error(f"Не удалось прочитать JSON-файл:\n{e}")
            return

        if not isinstance(data, list):
            self._show_error("Ожидается JSON-массив вопросов (как в А1.json / Б 2.4..json).")
            return

        try:
            wb = load_workbook(excel_file)
        except Exception as e:
            self._show_error(f"Не удалось открыть Excel-файл:\n{e}")
            return

        # Найти свободное имя листа (если такое уже есть — добавить (1), (2), ...)
        base_name = sheet_name
        current_name = base_name
        idx = 1
        existing = set(wb.sheetnames)
        while current_name in existing:
            current_name = f"{base_name} ({idx})"
            idx += 1

        ws = wb.create_sheet(title=current_name)

        # Заголовок
        headers = [
            "№ п.п.",
            "Вопрос",
            "1",
            "проверка",
            "2",
            "проверка2",
            "3",
            "проверка3",
            "4",
            "проверка4",
            "5",
            "проверка5",
            "6",
            "проверка6",
            "Нормативная основа вопроса",
        ]
        ws.append(headers)

        # Карта: номер варианта -> индекс столбца (начиная с 1)
        # A=1, B=2, ...
        option_col_index = {
            1: 3,   # "1"
            2: 5,   # "2"
            3: 7,   # "3"
            4: 9,   # "4"
            5: 11,  # "5"
            6: 13,  # "6"
        }
        check_col_index = {
            1: 4,   # "проверка"
            2: 6,   # "проверка2"
            3: 8,   # "проверка3"
            4: 10,  # "проверка4"
            5: 12,  # "проверка5"
            6: 14,  # "проверка6"
        }

        # Заполнение строк из JSON
        for item in data:
            if not isinstance(item, dict):
                continue

            number = item.get("number", "")
            question = item.get("question", "")
            options = item.get("options", []) or []
            correct_list = item.get("correct", []) or []
            reference = item.get("reference", "")

            # Убедимся, что correct_list — список строк
            if isinstance(correct_list, str):
                correct_list = [correct_list]

            # Базовые ячейки: №, вопрос, Нормативная основа
            row = ["" for _ in range(len(headers))]
            row[0] = number
            row[1] = question
            row[-1] = reference

            # Варианты и проверки
            for idx_opt, text in enumerate(options, start=1):
                if idx_opt > 6:
                    break
                col_opt = option_col_index[idx_opt] - 1  # в списке индекс с 0
                col_check = check_col_index[idx_opt] - 1
                row[col_opt] = text
                # "+" если этот текст есть среди правильных ответов
                row[col_check] = "+" if text in correct_list else "-"

            ws.append(row)

        # ---------- Оформление листа ----------
        # Общий шрифт и выравнивание + перенос строк
        default_font = Font(name="Arial", size=12)
        default_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        max_col = ws.max_column
        max_row = ws.max_row

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.font = default_font
                cell.alignment = default_alignment

        # Заголовок — жирный
        for cell in ws[1]:
            cell.font = Font(name="Arial", size=12, bold=True)
        # Высота первой строки
        ws.row_dimensions[1].height = 27.6

        # Колонка A, начиная со 2-й строки: Calibri 11
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=1)  # A
            cell.font = Font(name="Calibri", size=11)

        # Чередование цвета строк (каждая вторая, начиная со 2-й)
        alt_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for row_idx in range(2, max_row + 1):
            if row_idx % 2 == 0:
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = alt_fill

        # Раскраска плюсов/минусов в столбцах проверки (заливка, текст черный)
        plus_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        minus_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        check_columns = [4, 6, 8, 10, 12, 14]  # D, F, H, J, L, N
        for row_idx in range(2, max_row + 1):
            for col_idx in check_columns:
                cell = ws.cell(row=row_idx, column=col_idx)
                value = str(cell.value).strip() if cell.value is not None else ""
                if value == "+":
                    cell.fill = plus_fill
                elif value == "-":
                    cell.fill = minus_fill

        # Границы "Все границы" для всей таблицы
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border

        # Фиксированная ширина столбцов по ТЗ (в единицах Excel)
        fixed_widths = {
            "A": 8.11,
            "B": 58.89,
            "C": 50.78,
            "D": 12.89,
            "E": 47.11,
            "F": 13.56,
            "G": 45.78,
            "H": 13.56,
            "I": 51.11,
            "J": 13.56,
            "K": 40.33,
            "L": 13.56,
            "M": 35.78,
            "N": 14.67,
            "O": 51.67,
        }
        for col_letter, width in fixed_widths.items():
            ws.column_dimensions[col_letter].width = width

        try:
            wb.save(excel_file)
        except Exception as e:
            self._show_error(f"Не удалось сохранить Excel-файл:\n{e}")
            return

        self._show_info(f"Лист '{current_name}' успешно создан и заполнен.")


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
